from typing import Any

import openpyxl
from django.conf import settings
from django.contrib import messages
from django.contrib.admin.views.decorators import staff_member_required
from django.core.mail import EmailMessage
from django.db import IntegrityError
from django.http import HttpRequest, HttpResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.template.loader import render_to_string
from django.urls import reverse
from django.utils.decorators import method_decorator
from django.utils.html import strip_tags
from django.utils.timezone import now
from django.utils.translation import gettext_lazy as _
from django.views import View
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from blog.models import Partenaire

from .forms import DynamicEventForm
from .models import Evenement, InteretEvenement
from .utils.export_excel import export_all_to_excel


def liste_evenements(request):
    evenements = Evenement.objects.filter(est_actif=True).order_by("-date_debut")
    return render(request, "evenements/liste.html", {"evenements": evenements})


def detail_evenement(request, slug):
    evenement = get_object_or_404(Evenement, slug=slug)
    partenaires = Partenaire.objects.filter(archive=False)
    lang = request.LANGUAGE_CODE
    image_url = evenement.get_image_by_lang(lang) if evenement else None

    existing_registration = False
    if request.user.is_authenticated:
        existing_registration = InteretEvenement.objects.filter(
            utilisateur=request.user, evenement=evenement
        ).exists()
        partenaires = Partenaire.objects.filter(archive=False)
        lang = request.LANGUAGE_CODE
        image_url = evenement.get_image_by_lang(lang) if evenement else None

    return render(
        request,
        "evenements/detail.html",
        {
            "evenement": evenement,
            "existing_registration": existing_registration,
            "now": now(),
            "partenaires": partenaires,
            "image_url": image_url,
        },
    )


def envoyer_email_confirmation(request, inscription, evenement):
    detail_url = request.build_absolute_uri(
        reverse("evenements:detail", kwargs={"slug": evenement.slug})
    )
    sujet = _("Confirmation d'inscription à l'événement {}").format(evenement.titre)

    context = {
        "nom": inscription.nom_complet,
        "evenement": evenement,
        "date_debut": evenement.date_debut.strftime("%d/%m/%Y"),
        "date_fin": evenement.date_fin.strftime("%d/%m/%Y"),
        "detail_url": detail_url,
    }

    html_message = render_to_string("evenements/email_confirmation.html", context)
    plain_message = strip_tags(html_message)

    EmailMessage(
        subject=sujet,
        body=plain_message,
        from_email=settings.DEFAULT_FROM_EMAIL,
        to=[inscription.email],
        reply_to=[settings.DEFAULT_FROM_EMAIL],
    ).send()


def interet_evenement(request, slug):
    evenement = get_object_or_404(Evenement, slug=slug)
    questionnaire = evenement.questionnaire

    # Vérifier l'inscription existante
    existing_registration = None
    if request.user.is_authenticated:
        existing_registration = InteretEvenement.objects.filter(
            evenement=evenement, utilisateur=request.user
        ).first()
    elif request.method == "GET" and "email" in request.GET:
        existing_registration = InteretEvenement.objects.filter(
            evenement=evenement, email=request.GET["email"]
        ).first()

    # Préparer les données initiales
    initial_data = {}
    if existing_registration:
        initial_data = {
            "nom_complet": existing_registration.nom_complet,
            "email": existing_registration.email,
            "telephone": existing_registration.telephone,
            "interesse": existing_registration.interesse,
            "accepte_newsletter": existing_registration.accepte_newsletter,
        }

        for reponse in existing_registration.reponses.select_related("question"):
            field_name = f"question_{reponse.question.id}"
            if reponse.question.type_question == "CM":
                initial_data[field_name] = reponse.valeur.split(";")
            elif reponse.question.type_question == "BL":
                initial_data[field_name] = reponse.valeur.lower() in (
                    "true",
                    "1",
                    "oui",
                    "yes",
                )
            else:
                initial_data[field_name] = reponse.valeur

    # Gestion du formulaire
    if request.method == "POST":
        form = DynamicEventForm(
            request.POST, request.FILES, questionnaire=questionnaire
        )
        if form.is_valid():
            try:
                inscription = form.save(
                    evenement=evenement,
                    utilisateur=request.user if request.user.is_authenticated else None,
                )
                envoyer_email_confirmation(request, inscription, evenement)
                messages.success(
                    request,
                    _(
                        "Votre inscription a bien été enregistrée. Un email de confirmation vous a été envoyé."
                    ),
                )
                return redirect("evenements:detail", slug=slug)
            except IntegrityError:
                messages.error(
                    request, _("Une erreur est survenue lors de l'inscription")
                )
    else:
        form = DynamicEventForm(initial=initial_data, questionnaire=questionnaire)

    # Préparation des données pour le template
    sections_with_fields = []
    if questionnaire:
        for section in questionnaire.sections.all():
            section_data = {"section": section, "questions": []}
            for sq in section.sectionquestion_set.all():
                field_name = f"question_{sq.question.id}"
                if field_name in form.fields:
                    field_bound = form[field_name]
                    section_data["questions"].append(
                        {
                            "field": field_bound,
                            "question": sq.question,
                            "widget_type": field_bound.field.widget.__class__.__name__,  # <-- ajouté ici
                        }
                    )
            sections_with_fields.append(section_data)

    context = {
        "form": form,
        "evenement": evenement,
        "existing_registration": existing_registration,
        "sections_with_fields": sections_with_fields,
        "questionnaire": questionnaire,
        "questionnaire_image": (
            evenement.questionnaire.image_fr if evenement.questionnaire else None
        ),
    }

    return render(request, "evenements/interet.html", context)


@method_decorator(staff_member_required, name="dispatch")
class ExportEvenementExcelView(View):

    def get(self, request, evenement_id, *args, **kwargs):
        evenement = get_object_or_404(Evenement, id=evenement_id)
        inscriptions = InteretEvenement.objects.filter(evenement=evenement)

        # Récupérer les questions liées au questionnaire de l’événement
        questions = []
        if evenement.questionnaire:
            for section in evenement.questionnaire.sections.all():
                questions.extend(section.questions.all().order_by("ordre"))

        # Créer un fichier Excel
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Inscriptions"

        # En-têtes fixes
        headers = [
            "Nom complet",
            "Email",
            "Téléphone",
            "Intéressé",
            "Newsletter",
            "Date d'inscription",
        ]
        # En-têtes dynamiques (questions du questionnaire)
        headers += [q.libelle for q in questions]
        ws.append(headers)

        # Remplissage des données
        for inscription in inscriptions:
            row = [
                inscription.nom_complet,
                inscription.email,
                inscription.telephone,
                "Oui" if inscription.interesse else "Non",
                "Oui" if inscription.accepte_newsletter else "Non",
                inscription.date_creation.strftime("%d/%m/%Y %H:%M"),
            ]
            for question in questions:
                valeur = inscription.get_reponse_for_question(question)
                if valeur is None:
                    row.append("")
                elif question.type_question == "FL":  # Fichier
                    # si une réponse fichier existe → ajouter le lien ou le nom du fichier
                    reponse = inscription.reponses.filter(question=question).first()
                    row.append(
                        reponse.fichier.url if reponse and reponse.fichier else ""
                    )
                elif question.type_question == "BL":  # Booléen
                    row.append("Oui" if valeur else "Non")
                else:
                    row.append(valeur)
            ws.append(row)

        # Ajustement largeur colonnes
        for i, col in enumerate(ws.columns, start=1):
            max_length = 0
            column = get_column_letter(i)
            for cell in col:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            ws.column_dimensions[column].width = max_length + 2

        # Réponse HTTP pour téléchargement
        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = (
            f'attachment; filename="inscriptions_{evenement_id}.xlsx"'
        )
        wb.save(response)
        return response


@method_decorator(staff_member_required, name="dispatch")
class ExcelView(View):
    template_name = "evenements/export_excel.html"
    context: dict[str, Any] = {}

    def get(self, request, *args, **kwargs):

        evenements = Evenement.objects.filter(slug="b2b-turin-2026").first()
        inscriptions = InteretEvenement.objects.filter(evenement=evenements)

        return render(
            request,
            self.template_name,
            {"evenements": evenements, "inscriptions": inscriptions},
        )

    def post(self, request, *args, **kwargs):
        return HttpResponse("POST request!")


@method_decorator(staff_member_required, name="dispatch")
class ExcelimportView(View):
    template_name = "evenements/import_excel.html"
    context: dict[str, Any] = {}

    def get(self, request, *args, **kwargs):

        evenement = Evenement.objects.filter(slug="b2b-turin-2026").first()
        inscriptions = InteretEvenement.objects.filter(evenement=evenement)

        return render(
            request,
            self.template_name,
            {"evenement": evenement, "inscriptions": inscriptions},
        )

    def post(self, request, *args, **kwargs):
        return HttpResponse("POST request!")


@method_decorator(staff_member_required, name="dispatch")
class ImportEvenementExcelView(View):
    template_name = "evenements/import_excel.html"

    def get(self, request: HttpRequest, evenement_id: int, *args, **kwargs):
        evenement = get_object_or_404(Evenement, id=evenement_id)
        return render(request, self.template_name, {"evenement": evenement})

    def post(self, request: HttpRequest, evenement_id: int, *args, **kwargs):
        evenement = get_object_or_404(Evenement, id=evenement_id)
        file = request.FILES.get("fichier_excel")

        if not file:
            messages.error(request, "Aucun fichier sélectionné.")
            return redirect(request.path)

        # Validation basique du fichier (type + taille) avant traitement
        # openpyxl, pour eviter les zip-bombs / fichiers arbitraires.
        if not file.name.lower().endswith((".xlsx", ".xlsm")):
            messages.error(request, "Format invalide : fichier .xlsx attendu.")
            return redirect(request.path)
        if file.size > 5 * 1024 * 1024:
            messages.error(request, "Fichier trop volumineux (max 5 Mo).")
            return redirect(request.path)

        try:
            self.handle_import(file, evenement)
            messages.success(request, "Importation réussie !")
        except Exception:
            # Ne pas divulguer le detail de l'exception a l'utilisateur.
            messages.error(request, "Erreur lors de l'import du fichier.")

        return redirect(request.path)

    def handle_import(self, file, evenement: Evenement):
        wb = load_workbook(file)
        ws = wb.active

        # Récupérer les entêtes
        headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]

        # Vérifier les colonnes minimales
        colonnes_obligatoires = [
            "Nom complet",
            "Email",
            "Téléphone",
            "Intéressé",
            "Newsletter",
        ]
        for col in colonnes_obligatoires:
            if col not in headers:
                raise ValueError(f"Colonne obligatoire manquante : {col}")

        # Questions dynamiques liées au questionnaire
        questions = []
        if evenement.questionnaire:
            for section in evenement.questionnaire.sections.all():
                questions.extend(section.questions.all().order_by("ordre"))

        # Dictionnaire de correspondance {libellé_question: Question}
        question_map = {q.libelle: q for q in questions}

        # Parcourir les lignes
        for row in ws.iter_rows(min_row=2, values_only=True):
            data = dict(zip(headers, row))

            # Créer ou mettre à jour l’inscription
            inscription, created = InteretEvenement.objects.update_or_create(
                evenement=evenement,
                email=data.get("Email"),
                defaults={
                    "nom_complet": data.get("Nom complet") or "",
                    "telephone": data.get("Téléphone") or "",
                    "interesse": True if data.get("Intéressé") == "Oui" else False,
                    "accepte_newsletter": (
                        True if data.get("Newsletter") == "Oui" else False
                    ),
                },
            )

            # Enregistrer les réponses dynamiques
            for libelle, question in question_map.items():
                valeur = data.get(libelle)
                if valeur is None or valeur == "":
                    continue

                # Gestion selon le type de question
                if question.type_question == "BL":  # Booléen
                    valeur = (
                        True
                        if str(valeur).strip().lower() in ["oui", "true", "1"]
                        else False
                    )

                # Sauvegarder la réponse
                inscription.set_reponse_for_question(question, valeur)


class ExportExcelPageView(View):
    template_name = "evenements/export_page.html"
    context: dict[str, Any] = {}

    def get(self, request, *args, **kwargs):

        return render(request, self.template_name)

    def post(self, request, *args, **kwargs):
        return HttpResponse("POST request!")


@method_decorator(staff_member_required, name="dispatch")
class ExportExcelDownloadView(View):
    """Génère et télécharge le fichier Excel."""

    def get(self, request):
        return export_all_to_excel()
