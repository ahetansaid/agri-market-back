from collections import Counter
from datetime import datetime

import openpyxl
from django.http import HttpResponse
from openpyxl.utils import get_column_letter

from accounts.models import Societe, Utilisateur
from evenements.models import InteretEvenement, Question, ReponseQuestion


def export_all_to_excel():
    wb = openpyxl.Workbook()

    # === 1️⃣ Feuille : Utilisateurs ===
    ws1 = wb.active
    ws1.title = "Utilisateurs"

    utilisateurs = Utilisateur.objects.all().values(
        "id",
        "username",
        "email",
        "telephone",
        "pays",
        "ville",
        "adresse",
        "code_postal",
        "user_type",
        "first_name",
        "last_name",
        "individual_category",
        "created_at",
    )

    headers = [
        "ID",
        "Nom d'utilisateur",
        "Email",
        "Téléphone",
        "Pays",
        "Ville",
        "Adresse",
        "Code postal",
        "Type d'utilisateur",
        "Prénom",
        "Nom",
        "Catégorie d'individu",
        "Date de création",
    ]
    ws1.append(headers)

    for u in utilisateurs:
        created_at = u.get("created_at")
        if isinstance(created_at, datetime) and created_at.tzinfo:
            created_at = created_at.replace(tzinfo=None)
        u["created_at"] = created_at.strftime("%Y-%m-%d %H:%M") if created_at else ""

        ws1.append(
            [
                u["id"],
                u["username"],
                u["email"],
                str(u["telephone"] or ""),
                u["pays"],
                u["ville"],
                u["adresse"],
                u["code_postal"],
                u["user_type"],
                u["first_name"],
                u["last_name"],
                u["individual_category"],
                u["created_at"],
            ]
        )

    # === 2️⃣ Feuille : Sociétés ===
    ws2 = wb.create_sheet(title="Sociétés")
    ws2_headers = [
        "ID",
        "Nom société",
        "Type d'entreprise",
        "Autre type",
        "Code commercial",
        "Utilisateur",
        "Email utilisateur",
        "Produits vendus",
        "Produits recherchés",
    ]
    ws2.append(ws2_headers)

    societes = (
        Societe.objects.select_related("utilisateur")
        .prefetch_related("produits_vendus", "produits_recherches")
        .all()
    )

    for s in societes:
        produits_vendus = ", ".join([p.name for p in s.produits_vendus.all()])
        produits_recherches = ", ".join([p.name for p in s.produits_recherches.all()])
        ws2.append(
            [
                s.id,
                s.nom,
                s.company_type,
                s.autre_type or "",
                s.code_commercial or "",
                s.utilisateur.username if s.utilisateur else "",
                s.utilisateur.email if s.utilisateur else "",
                produits_vendus,
                produits_recherches,
            ]
        )

    # === 3️⃣ Feuille : Inscriptions Événements ===
    ws3 = wb.create_sheet(title="Inscriptions Événements")
    ws3_headers = [
        "ID",
        "Nom complet",
        "Email",
        "Téléphone",
        "Intéressé",
        "Confirmé",
        "Événement",
        "Utilisateur",
        "Email utilisateur",
        "Accepte newsletter",
        "Date d'inscription",
    ]
    ws3.append(ws3_headers)

    interets = InteretEvenement.objects.select_related("utilisateur", "evenement").all()

    for i in interets:
        date_inscription = i.date_creation
        if isinstance(date_inscription, datetime) and date_inscription.tzinfo:
            date_inscription = date_inscription.replace(tzinfo=None)
        date_str = (
            date_inscription.strftime("%Y-%m-%d %H:%M") if date_inscription else ""
        )

        ws3.append(
            [
                i.id,
                i.nom_complet,
                i.email,
                str(i.telephone) if i.telephone else "",
                "Oui" if i.interesse else "Non",
                "Oui" if getattr(i, "confirme", False) else "Non",
                i.evenement.titre if i.evenement else "",
                i.utilisateur.username if i.utilisateur else "",
                i.utilisateur.email if i.utilisateur else "",
                "Oui" if i.accepte_newsletter else "Non",
                date_str,
            ]
        )

    # === 4️⃣ Feuille : Vue Globale ===
    ws4 = wb.create_sheet(title="Vue Globale")
    ws4_headers = [
        "Nom d'utilisateur",
        "Email",
        "Téléphone",
        "Type d'utilisateur",
        "Pays",
        "Ville",
        "Nom société",
        "Type d'entreprise",
        "Produits vendus",
        "Produits recherchés",
        "Événements inscrits",
        "Date de création du compte",
    ]
    ws4.append(ws4_headers)

    for user in Utilisateur.objects.all().prefetch_related("societe"):
        societe = getattr(user, "societe", None)
        interets_user = InteretEvenement.objects.filter(
            utilisateur=user
        ).select_related("evenement")

        evenements = ", ".join(
            [i.evenement.titre for i in interets_user if i.evenement]
        )
        produits_vendus = (
            ", ".join([p.name for p in societe.produits_vendus.all()])
            if societe
            else ""
        )
        produits_recherches = (
            ", ".join([p.name for p in societe.produits_recherches.all()])
            if societe
            else ""
        )

        created_at = user.created_at
        if isinstance(created_at, datetime) and created_at.tzinfo:
            created_at = created_at.replace(tzinfo=None)
        created_str = created_at.strftime("%Y-%m-%d %H:%M") if created_at else ""

        ws4.append(
            [
                user.username,
                user.email,
                str(user.telephone or ""),
                user.user_type,
                str(user.pays or ""),
                user.ville or "",
                societe.nom if societe else "",
                societe.company_type if societe else "",
                produits_vendus,
                produits_recherches,
                evenements,
                created_str,
            ]
        )

    # === 5️⃣ Feuille : Statistiques ===
    ws5 = wb.create_sheet(title="Statistiques")

    total_utilisateurs = Utilisateur.objects.count()
    total_societes = Societe.objects.count()
    total_interets = InteretEvenement.objects.count()
    total_interesses = InteretEvenement.objects.filter(interesse=True).count()

    # 🔍 Nouvelle partie : calcul du nombre de participants confirmés à partir des réponses

    # Récupérer la question de confirmation (ex: "Confirmez-vous votre participation ?")
    question_confirme = Question.objects.filter(libelle__icontains="confirme").first()

    if question_confirme:
        reponses_confirmees = ReponseQuestion.objects.filter(
            question=question_confirme, valeur__iregex=r"^(oui|confirm|ok|yes)$"
        ).select_related("inscription__utilisateur")

        # interets_confirmes = [r.inscription for r in reponses_confirmees]
        utilisateurs_confirmes = [
            r.inscription.utilisateur
            for r in reponses_confirmees
            if r.inscription.utilisateur
        ]
        total_confirmes = len(utilisateurs_confirmes)

    else:
        # interets_confirmes = []
        utilisateurs_confirmes = []
        total_confirmes = 0

    # 📊 Statistiques générales
    user_type_counts = Counter(Utilisateur.objects.values_list("user_type", flat=True))

    ws5.append(["Statistique", "Valeur"])
    ws5.append(["Nombre total d'utilisateurs", total_utilisateurs])
    for user_type, count in user_type_counts.items():
        ws5.append([f"↳ {user_type}", count])
        ws5.append(["Nombre total de sociétés", total_societes])
        ws5.append(["Nombre total d'inscriptions aux événements", total_interets])
        ws5.append(["Nombre total d'intéressés", total_interesses])
        ws5.append(["Nombre total de confirmations", total_confirmes])

        ws5.append([])
        ws5.append(["Légende :"])
        ws5.append(
            ["- 'Intéressé' = Utilisateur ayant marqué un intérêt pour un événement"]
        )
        ws5.append(["- 'Confirmé' = Utilisateur ayant confirmé sa participation"])
        ws5.append(
            ["- Les utilisateurs peuvent être liés à une société ou être indépendants"]
        )

    # === Ajustement automatique des colonnes ===
    for ws in [ws1, ws2, ws3, ws4, ws5]:
        for col in ws.columns:
            max_length = 0
            column = get_column_letter(col[0].column)
            for cell in col:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except Exception:
                    pass
            ws.column_dimensions[column].width = max_length + 3

    # === Génération du fichier Excel ===
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="export_complet.xlsx"'
    wb.save(response)
    return response
